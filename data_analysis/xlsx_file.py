from openpyxl import load_workbook

wb = load_workbook('REVIEWS.xlsx')
print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Form Responses 1')

# Print the sheet title 
sheet.title

# Get currently active sheet
anotherSheet = wb.active

d ={}
for i in range(2,sheet.max_row):
	if sheet['B'+str(i)].value in d:
		d[sheet['B'+str(i)].value] = (d[sheet['B'+str(i)].value] + (int(sheet['C'+str(i)].value)+int(sheet['D'+str(i)].value)+int(sheet['E'+str(i)].value))/3)/2
	else:
		d[sheet['B'+str(i)].value] = (int(sheet['C'+str(i)].value)+int(sheet['D'+str(i)].value)+int(sheet['E'+str(i)].value))/3

print d

