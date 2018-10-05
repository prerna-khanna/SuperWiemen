import pyrebase
from geopy.geocoders import Nominatim
import csv
from math import radians, cos, sin, asin, sqrt

from googlesearch import search
import requests
from bs4 import BeautifulSoup
import matplotlib.pyplot as plt
from textblob import TextBlob
import sys  
import os
from openpyxl import load_workbook
reload(sys)  
sys.setdefaultencoding('utf8')


def crowd_sourced(q):
	q.replace("delhi",'')
	wb = load_workbook('REVIEWS.xlsx')

	sheet = wb.get_sheet_by_name('Form Responses 1')

	# Print the sheet title 
	sheet.title

	# Get currently active sheet
	anotherSheet = wb.active

	d ={}
	for i in range(2,sheet.max_row):
		if sheet['B'+str(i)].value in d:
			d[sheet['B'+str(i)].value] = (d[sheet['B'+str(i)].value] + (int(sheet['C'+str(i)].value)+int(sheet['D'+str(i)].value)+int(sheet['E'+str(i)].value))/3)/2
		else:
			d[sheet['B'+str(i)].value] = (int(sheet['C'+str(i)].value)+int(sheet['D'+str(i)].value)+int(sheet['E'+str(i)].value))/3

	print d
	if q in d:
		return d[q]
	else:
		os.system("python query.py "+ q)
		fh = open("hello.txt","r")
		return float(fh.read())
	

def stream_handler(message):
  
  
   entry_len = len(message["path"])
   entry = message["path"]
   
   if(message and message['data'] != None):

        home = db.child("%s"%message["path"]).get().val()

        if isinstance(home, dict):
            print("true")
        else:
            user = entry.split('/')[2]
            
            dest = db.child("uploads").child(user).child("dest").get().val()
            work = db.child("uploads").child(user).child("work").get().val()
            unique = db.child("uploads").child(user).child("unique").get().val()

            print(dest)
	    r1 = crowd_sourced(dest)
	    print r1

            geolocator = Nominatim(user_agent="specify_your_app_name_here")
            location = geolocator.geocode(dest, timeout=None)

            lat1 = location.latitude
            lon1 = location.longitude
	    db.child("uploads").child(user).child("latitude").set(lat1)
	    db.child("uploads").child(user).child("longitude").set(lon1)


            """
            Calculate the great circle distance between two points 
            on the earth (specified in decimal degrees)
            """
            csvfile = open('crime_csv_1.csv')
            csv_reader = csv.reader(csvfile)
            next(csv_reader)
            crime_score = 0
            radius = 5.00
            mean = 0
            for row in csv_reader:

                mean = mean+int(row[3])

                lat2 = float(row[1])
                lon2 = float(row[2])

                # convert decimal degrees to radians 
                # lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])
                

                # haversine formula 
                dlon = radians(lon2 - lon1)
                dlat = radians(lat2 - lat1)
                a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
                c = 2 * asin(sqrt(a)) 
                r = 6371 # Radius of earth in kilometers. Use 3956 for miles
                result = c * r
               
                if (result<=radius):
                    crime_score += int(row[3])

            crime_score_normalised = (abs(crime_score-mean)/mean) * 5

            r2 = 5 - crime_score_normalised
	    print r2
	    r = int((r1+r2)/2)
	    print "final rating " + str(r)
	    db.child("uploads").child(user).child("rate").set(str(r))
	
	    


config = {
    "apiKey": " AIzaSyDStzGJEAUy-Js1vnW2IygbiXl43cTvElI ",
    "authDomain": "superwiemenfirebaseapp.com",
    "databaseURL": "https://superwiemen.firebaseio.com/",
    "storageBucket": "superwiemen.appspot.com",
    #"serviceAccount": "path/to/serviceAccountCredentials.json"
    }

firebase = pyrebase.initialize_app(config)
db = firebase.database()

st = firebase.storage()

my_stream = db.stream(stream_handler, stream_id="new_posts")
